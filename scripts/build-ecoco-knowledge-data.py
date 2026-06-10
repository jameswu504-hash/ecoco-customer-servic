import argparse
import json
import zipfile
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


def clean(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    while "\n\n\n" in text:
        text = text.replace("\n\n\n", "\n\n")
    return text


def build(input_dir, output_dir):
    input_dir = Path(input_dir)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(input_dir / "faq.xlsx", data_only=True, read_only=True)
    sections = []

    with zipfile.ZipFile(input_dir / "command_center.zip") as archive:
        with archive.open("ECOCO_CS_CommandCenter_v1.9.3/config.json") as handle:
            config = json.loads(handle.read().decode("utf-8"))
            brand_context = clean(config.get("brand_context"))

    sections.append(
        {
            "category": "CommandCenter 品牌語氣與客服原則",
            "content": (
                brand_context
                + "\n\n注意：本段只抽取 brand_context，不包含舊版 api_key、zd_token 或 mail 設定。"
            ),
            "source": "ECOCO_CS_CommandCenter_v1.9.3/config.json:brand_context",
            "visibility": "public_answer_rules",
        }
    )

    meta_text = clean((input_dir / "meta_prompt.md").read_text(encoding="utf-8"))
    sections.append(
        {
            "category": "Meta 社群客服回覆規則",
            "content": meta_text,
            "source": "目前給Meta ai 指令.md",
            "visibility": "response_rules",
        }
    )

    latest_faq = workbook.worksheets[2]
    faq_lines = ["以下內容來自最新版官網常見問題 20260515，適合作為客戶可回答知識。"]
    for row in list(latest_faq.iter_rows(values_only=True))[1:]:
        question = clean(row[0] if len(row) > 0 else "")
        answer = clean(row[2] if len(row) > 2 else "")
        link = clean(row[3] if len(row) > 3 else "")
        if not question or not answer or question.startswith("#") or answer.startswith("#"):
            continue
        faq_lines.append(f"\n### {question}\n{answer}")
        if link:
            faq_lines.append(f"連結：{link}")
    sections.append(
        {
            "category": "官網常見問題 20260515",
            "content": "\n".join(faq_lines).strip(),
            "source": "凡立橙股份有限公司_官網常見問題_茗芬V2 的副本.xlsx / 官網常見問題 20260515",
            "visibility": "public_knowledge",
        }
    )

    reply_sheet = workbook.worksheets[0]
    groups = defaultdict(list)
    for row in list(reply_sheet.iter_rows(values_only=True))[1:]:
        topic = clean(row[0] if len(row) > 0 else "") or "未分類"
        main = clean(row[1] if len(row) > 1 else "")
        sub = clean(row[2] if len(row) > 2 else "")
        detail = clean(row[3] if len(row) > 3 else "")
        content = clean(row[5] if len(row) > 5 else "") or clean(row[4] if len(row) > 4 else "")
        created = clean(row[6] if len(row) > 6 else "")
        adjusted = clean(row[7] if len(row) > 7 else "")
        if not content or content.startswith("#"):
            continue

        title_parts = [part for part in [main, sub, detail] if part]
        title = " / ".join(title_parts) if title_parts else topic
        dates = []
        if created:
            dates.append(f"建立：{created}")
        if adjusted:
            dates.append(f"調整：{adjusted}")
        date_line = f"\n日期：{'；'.join(dates)}" if dates else ""
        groups[topic].append(f"### {title}{date_line}\n{content}")

    for topic in sorted(groups):
        sections.append(
            {
                "category": f"客服回覆問答：{topic}",
                "content": "\n\n".join(groups[topic]),
                "source": "凡立橙股份有限公司_官網常見問題_茗芬V2 的副本.xlsx / 回覆問答",
                "visibility": "public_knowledge_or_agent_assist",
            }
        )

    internal_sheet = workbook.worksheets[1]
    internal_lines = ["內部使用：這些是客服/營運觸發備註，不建議對客戶逐字揭露。"]
    for row in list(internal_sheet.iter_rows(values_only=True))[1:]:
        note = clean(row[0] if len(row) > 0 else "")
        keyword = clean(row[1] if len(row) > 1 else "")
        modifier = clean(row[2] if len(row) > 2 else "")
        ops = clean(row[3] if len(row) > 3 else "")
        line_parts = []
        if note:
            line_parts.append(f"內部備註：{note}")
        if keyword:
            line_parts.append(f"關鍵字：{keyword}")
        if modifier:
            line_parts.append(f"修飾用語：{modifier}")
        if ops:
            line_parts.append(f"營運用語：{ops}")
        if line_parts:
            internal_lines.append("- " + "；".join(line_parts))
    sections.append(
        {
            "category": "內部關鍵字與營運觸發備註",
            "content": "\n".join(internal_lines),
            "source": "凡立橙股份有限公司_官網常見問題_茗芬V2 的副本.xlsx / 關鍵字_內部備註",
            "visibility": "internal_agent_assist",
        }
    )

    policies = [
        {
            "key": "points_missing",
            "name": "點數未入帳",
            "risk": "medium",
            "automation_level": "draft_review",
            "required_fields": [
                "註冊手機",
                "使用日期時間",
                "站點名稱",
                "投入品項",
                "投入數量",
                "成功或錯誤畫面截圖",
            ],
            "do_not_say": ["保證補點", "保證立即處理完成", "自行猜測原因"],
            "action": "蒐集資訊後轉客服或後端查詢",
        },
        {
            "key": "coupon_issue",
            "name": "優惠券兌換或使用異常",
            "risk": "high",
            "automation_level": "draft_review",
            "required_fields": ["會員帳號", "兌換日期", "券名稱", "使用店家", "消費截圖", "錯誤畫面"],
            "do_not_say": ["直接承諾退點", "直接承諾補發優惠券"],
            "action": "蒐集證明後由人工審核",
        },
        {
            "key": "machine_issue",
            "name": "機台異常或滿袋",
            "risk": "medium",
            "automation_level": "draft_review",
            "required_fields": ["站點名稱", "機台類型", "發生時間", "異常描述", "照片或影片"],
            "do_not_say": ["保證維修時間", "指責使用者操作錯誤"],
            "action": "先同理，再提供 App 查詢或客服表單回報方式",
        },
        {
            "key": "simple_faq",
            "name": "一般 FAQ",
            "risk": "low",
            "automation_level": "auto_reply_allowed",
            "required_fields": [],
            "do_not_say": ["超出知識庫自行推測"],
            "action": "可依知識庫直接回答",
        },
    ]

    generated_at = datetime.now().isoformat(timespec="seconds")
    knowledge_payload = {
        "generated_at": generated_at,
        "notes": "Generated from ECOCO files. Secrets from old config.json are intentionally excluded.",
        "source_files": [
            "ECOCO_CS_CommandCenter_v1.9.3/config.json brand_context only",
            "目前給Meta ai 指令.md",
            "凡立橙股份有限公司_官網常見問題_茗芬V2 的副本.xlsx",
        ],
        "sections": sections,
    }

    policy_payload = {
        "generated_at": generated_at,
        "policies": policies,
    }

    (output_dir / "ecoco-knowledge-import.json").write_text(
        json.dumps(knowledge_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (output_dir / "ecoco-response-policies.json").write_text(
        json.dumps(policy_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(f"sections={len(sections)}")
    print(f"topic_sections={len(groups)}")
    print(f"policies={len(policies)}")


def main():
    parser = argparse.ArgumentParser(description="Build ECOCO knowledge import JSON from source files.")
    parser.add_argument("--input-dir", required=True, help="Directory containing faq.xlsx, meta_prompt.md, command_center.zip")
    parser.add_argument("--output-dir", default="data", help="Output directory for generated JSON files")
    args = parser.parse_args()
    build(args.input_dir, args.output_dir)


if __name__ == "__main__":
    main()

